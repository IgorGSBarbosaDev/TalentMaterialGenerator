from __future__ import annotations

from pathlib import Path

import pytest

from app.core.reader import (
    detect_columns,
    normalize_filename,
    parse_multiline_field,
    read_spreadsheet,
    validate_required_columns,
)


def _fixture_path(filename: str) -> Path:
    return Path(__file__).resolve().parents[2] / "fixtures" / filename


class TestReadSpreadsheet:
    def test_read_spreadsheet_returns_list_of_dicts(self) -> None:
        result = read_spreadsheet(str(_fixture_path("colaboradores_sample.xlsx")))

        assert isinstance(result, list)
        assert len(result) > 0
        assert isinstance(result[0], dict)

    def test_read_spreadsheet_row_count_matches_fixture(self) -> None:
        result = read_spreadsheet(str(_fixture_path("colaboradores_sample.xlsx")))

        assert len(result) == 3

    def test_read_spreadsheet_raises_file_not_found_for_missing_file(self) -> None:
        with pytest.raises(FileNotFoundError):
            read_spreadsheet(str(_fixture_path("arquivo_inexistente.xlsx")))

    def test_read_spreadsheet_empty_cells_return_empty_string_not_none(self) -> None:
        result = read_spreadsheet(str(_fixture_path("colaboradores_sample.xlsx")))

        assert result[2]["nota"] == ""
        for row in result:
            for value in row.values():
                assert value is not None

    def test_read_spreadsheet_empty_sheet_returns_empty_list(self) -> None:
        result = read_spreadsheet(str(_fixture_path("colaboradores_empty.xlsx")))

        assert result == []

    def test_read_spreadsheet_returns_empty_list_when_sheet_has_no_rows(
        self,
        tmp_path: Path,
        monkeypatch,
    ) -> None:
        class _FakeWorksheet:
            def iter_rows(self, values_only: bool = True):
                return iter(())

        class _FakeWorkbook:
            active = _FakeWorksheet()

        existing_file = tmp_path / "empty.xlsx"
        existing_file.write_bytes(b"placeholder")

        monkeypatch.setattr(
            "app.core.reader.load_workbook", lambda **kwargs: _FakeWorkbook()
        )

        assert read_spreadsheet(str(existing_file)) == []

    def test_read_spreadsheet_returns_empty_list_when_active_sheet_is_none(
        self,
        tmp_path: Path,
        monkeypatch,
    ) -> None:
        class _FakeWorkbook:
            active = None

        existing_file = tmp_path / "empty.xlsx"
        existing_file.write_bytes(b"placeholder")

        monkeypatch.setattr(
            "app.core.reader.load_workbook", lambda **kwargs: _FakeWorkbook()
        )

        assert read_spreadsheet(str(existing_file)) == []

    def test_read_spreadsheet_returns_empty_list_when_header_row_is_blank(
        self,
        tmp_path: Path,
    ) -> None:
        from openpyxl import Workbook

        workbook_path = tmp_path / "blank_headers.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.append(["   ", "  "])
        worksheet.append(["valor 1", "valor 2"])
        workbook.save(workbook_path)

        assert read_spreadsheet(str(workbook_path)) == []

    def test_read_spreadsheet_ignores_empty_header_and_blank_data_row(
        self,
        tmp_path: Path,
    ) -> None:
        from openpyxl import Workbook

        workbook_path = tmp_path / "mixed_headers.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.append(["nome", "   "])
        worksheet.append(["Ana Martins", "coluna ignorada"])
        worksheet.append([None, None])
        workbook.save(workbook_path)

        result = read_spreadsheet(str(workbook_path))

        assert result == [{"nome": "Ana Martins"}]


class TestDetectColumns:
    @pytest.mark.parametrize(
        ("header", "expected_field"),
        [
            ("nome", "nome"),
            ("name", "nome"),
            ("nome_completo", "nome"),
            ("colaborador", "nome"),
            ("funcionario", "nome"),
            ("idade", "idade"),
            ("age", "idade"),
            ("anos", "idade"),
            ("cargo", "cargo"),
            ("funcao", "cargo"),
            ("função", "cargo"),
            ("role", "cargo"),
            ("posição", "cargo"),
            ("posicao", "cargo"),
            ("antiguidade", "antiguidade"),
            ("tempo_empresa", "antiguidade"),
            ("anos_empresa", "antiguidade"),
            ("admissao", "antiguidade"),
            ("formacao", "formacao"),
            ("formação", "formacao"),
            ("graduacao", "formacao"),
            ("escolaridade", "formacao"),
            ("education", "formacao"),
            ("resumo", "resumo_perfil"),
            ("perfil", "resumo_perfil"),
            ("resumo_perfil", "resumo_perfil"),
            ("descricao", "resumo_perfil"),
            ("bio", "resumo_perfil"),
            ("trajetoria", "trajetoria"),
            ("trajetória", "trajetoria"),
            ("historico", "trajetoria"),
            ("histórico", "trajetoria"),
            ("carreira", "trajetoria"),
            ("performance", "performance"),
            ("avaliacao", "performance"),
            ("avaliação", "performance"),
            ("resultado", "performance"),
            ("nota_historico", "performance"),
            ("foto", "foto"),
            ("photo", "foto"),
            ("imagem", "foto"),
            ("image", "foto"),
            ("arquivo_foto", "foto"),
            ("area", "area"),
            ("área", "area"),
            ("departamento", "area"),
            ("setor", "area"),
            ("gerencia", "area"),
            ("potencial", "potencial"),
            ("potential", "potencial"),
            ("nota", "nota"),
            ("score", "nota"),
            ("avaliacao_atual", "nota"),
            ("resultado_atual", "nota"),
        ],
    )
    def test_detect_columns_maps_documented_variations(
        self,
        header: str,
        expected_field: str,
    ) -> None:
        mapping = detect_columns([header])

        assert mapping[expected_field] == header

    def test_detect_columns_maps_nome_exact(self) -> None:
        mapping = detect_columns(["nome"])

        assert mapping["nome"] == "nome"

    def test_detect_columns_maps_nome_completo_variant(self) -> None:
        mapping = detect_columns(["nome_completo"])

        assert mapping["nome"] == "nome_completo"

    def test_detect_columns_maps_colaborador_variant(self) -> None:
        mapping = detect_columns(["colaborador"])

        assert mapping["nome"] == "colaborador"

    def test_detect_columns_maps_cargo_exact(self) -> None:
        mapping = detect_columns(["cargo"])

        assert mapping["cargo"] == "cargo"

    def test_detect_columns_maps_funcao_variant(self) -> None:
        mapping = detect_columns(["funcao"])

        assert mapping["cargo"] == "funcao"

    def test_detect_columns_is_case_insensitive(self) -> None:
        mapping = detect_columns(["NoMe", "FuNcAo"])

        assert mapping["nome"] == "NoMe"
        assert mapping["cargo"] == "FuNcAo"

    def test_detect_columns_returns_none_for_unknown_header(self) -> None:
        mapping = detect_columns(["coluna_desconhecida_xyz"])

        assert all(value is None for value in mapping.values())

    def test_detect_columns_handles_accented_headers(self) -> None:
        mapping = detect_columns(["FUNÇÃO", "FORMAÇÃO", "ÁREA", "TRAJETÓRIA"])

        assert mapping["cargo"] == "FUNÇÃO"
        assert mapping["formacao"] == "FORMAÇÃO"
        assert mapping["area"] == "ÁREA"
        assert mapping["trajetoria"] == "TRAJETÓRIA"


class TestValidateRequiredColumns:
    def test_validate_required_columns_returns_empty_when_all_mapped(self) -> None:
        missing = validate_required_columns({"nome": "nome", "cargo": "cargo"})

        assert missing == []

    def test_validate_required_columns_returns_missing_nome(self) -> None:
        missing = validate_required_columns({"nome": None, "cargo": "cargo"})

        assert missing == ["nome"]

    def test_validate_required_columns_returns_missing_cargo(self) -> None:
        missing = validate_required_columns({"nome": "nome", "cargo": None})

        assert missing == ["cargo"]


class TestParseMultilineField:
    def test_parse_multiline_splits_by_semicolon(self) -> None:
        result = parse_multiline_field("Cargo A;Cargo B")

        assert result == ["Cargo A", "Cargo B"]

    def test_parse_multiline_splits_by_newline(self) -> None:
        result = parse_multiline_field("Cargo A\nCargo B")

        assert result == ["Cargo A", "Cargo B"]

    def test_parse_multiline_trims_whitespace_from_items(self) -> None:
        result = parse_multiline_field("  Cargo A ;   Cargo B  ")

        assert result == ["Cargo A", "Cargo B"]

    def test_parse_multiline_removes_empty_items(self) -> None:
        result = parse_multiline_field("Cargo A;;\n;Cargo B;\n")

        assert result == ["Cargo A", "Cargo B"]

    def test_parse_multiline_empty_string_returns_empty_list(self) -> None:
        assert parse_multiline_field("") == []


class TestNormalizeFilename:
    def test_normalize_replaces_spaces_with_underscores(self) -> None:
        assert normalize_filename("Ana Martins") == "Ana_Martins"

    def test_normalize_removes_accents(self) -> None:
        assert normalize_filename("João") == "Joao"

    def test_normalize_handles_empty_string(self) -> None:
        assert normalize_filename("") == ""

    def test_normalize_combined_accents_and_spaces(self) -> None:
        assert normalize_filename("Márcia Ávila") == "Marcia_Avila"
