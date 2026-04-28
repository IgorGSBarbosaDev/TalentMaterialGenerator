from __future__ import annotations

import hashlib
import os
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Final, TypedDict, cast

import requests
from openpyxl import load_workbook

from app.config import settings
from app.core.carom_templates import get_carom_preset

COLUMN_VARIATIONS: dict[str, tuple[str, ...]] = {
    "matricula": (
        "matricula",
        "matricula_funcional",
        "matricula_colaborador",
        "id",
        "id_unico",
    ),
    "nome": ("nome", "name", "nome_completo", "colaborador", "funcionario"),
    "idade": ("idade", "age", "anos"),
    "cargo": ("cargo", "funcao", "funcao_atual", "role", "posicao"),
    "antiguidade": ("antiguidade", "tempo_empresa", "anos_empresa", "admissao"),
    "formacao": ("formacao", "graduacao", "escolaridade", "education"),
    "resumo_perfil": (
        "resumo",
        "perfil",
        "resumo_perfil",
        "resumo_do_perfil",
        "descricao",
        "bio",
    ),
    "trajetoria": ("trajetoria", "historico", "carreira"),
    "nota_2025": ("nota_2025", "nota 2025", "avaliacao_2025", "avaliacao 2025"),
    "nota_2024": ("nota_2024", "nota 2024", "avaliacao_2024", "avaliacao 2024"),
    "nota_2023": ("nota_2023", "nota 2023", "avaliacao_2023", "avaliacao 2023"),
    "performance": (
        "performance",
        "avaliacao",
        "resultado",
        "nota_historico",
    ),
    "area": ("area", "departamento", "setor", "gerencia"),
    "potencial": ("potencial", "potential"),
    "nota": ("nota", "score", "avaliacao_atual", "resultado_atual"),
    "foto": ("foto", "photo", "photo_ref", "photo_reference", "imagem", "avatar"),
    "localizacao": ("localizacao", "location", "cidade", "site"),
    "unidade_gestao": ("unidade_gestao", "management_unit", "gerencia_unidade"),
    "ceo1": ("ceo1", "ceo_1"),
    "ceo2": ("ceo2", "ceo_2"),
    "ceo3": ("ceo3", "ceo_3"),
    "ceo4": ("ceo4", "ceo_4"),
}


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.strip().lower().replace(" ", "_")


NORMALIZED_VARIATIONS: dict[str, set[str]] = {
    field: {_normalize_text(item) for item in variations}
    for field, variations in COLUMN_VARIATIONS.items()
}

FICHA_EVALUATION_YEARS: Final[tuple[str, ...]] = ("2025", "2024", "2023")
FICHA_BASE_FIELDS: Final[tuple[str, ...]] = (
    "matricula",
    "nome",
    "idade",
    "cargo",
    "antiguidade",
    "formacao",
    "resumo_perfil",
    "trajetoria",
)
FICHA_FIELDS: Final[tuple[str, ...]] = (
    *FICHA_BASE_FIELDS,
    "nota_2025",
    "nota_2024",
    "nota_2023",
    "avaliacao_2025",
    "avaliacao_2024",
    "avaliacao_2023",
    "score_2025",
    "score_2024",
    "score_2023",
    "potencial_2025",
    "potencial_2024",
    "potencial_2023",
)
FICHA_REQUIRED_FIELDS: Final[tuple[str, ...]] = ("matricula", "nome", "cargo")
MAX_FICHA_NAME_MATCHES: Final = 25
EXPECTED_FICHA_COLUMN_ORDERS: Final[tuple[tuple[str, ...], ...]] = (
    (
        "Matricula",
        "Nome",
        "Idade",
        "Cargo",
        "Antiguidade",
        "Formacao",
        "Resumo do perfil",
        "Trajetoria",
        "Nota 2025",
        "Nota 2024",
        "Nota 2023",
    ),
    (
        "Matricula",
        "Nome",
        "Cargo",
        "Idade",
        "Antiguidade",
        "Formacao",
        "Resumo do perfil",
        "Trajetoria",
        "Avaliação 2025",
        "Avaliação 2024",
        "Avaliação 2023",
        "Nota 2025",
        "Potencial 2025",
        "Nota 2024",
        "Potencial 2024",
        "Nota 2023",
        "Potencial 2023",
    ),
)
CAROM_FIELDS: Final[tuple[str, ...]] = (
    "matricula",
    "nome",
    "idade",
    "cargo",
    "formacao",
    "resumo_perfil",
    "trajetoria",
    "foto",
    "area",
    "localizacao",
    "unidade_gestao",
    "ceo1",
    "ceo2",
    "nota_2025",
    "avaliacao_2025",
    "score_2025",
    "potencial_2025",
    "ceo3",
    "ceo4",
)
CAROM_REQUIRED_FIELDS: Final[tuple[str, ...]] = ("matricula", "nome", "cargo")


@dataclass
class SpreadsheetSourceResult:
    path: str
    source_kind: str
    is_temporary: bool
    used_cache: bool
    cache_path: str | None = None
    message: str = ""
    downloaded_at: str = ""


class FichaEmployee(TypedDict):
    matricula: str
    nome: str
    idade: str
    cargo: str
    antiguidade: str
    formacao: str
    resumo_perfil: str
    trajetoria: str
    nota_2025: str
    nota_2024: str
    nota_2023: str
    avaliacao_2025: str
    avaliacao_2024: str
    avaliacao_2023: str
    score_2025: str
    score_2024: str
    score_2023: str
    potencial_2025: str
    potencial_2024: str
    potencial_2023: str


class CaromEmployee(TypedDict):
    matricula: str
    nome: str
    idade: str
    cargo: str
    formacao: str
    resumo_perfil: str
    trajetoria: str
    foto: str
    area: str
    localizacao: str
    unidade_gestao: str
    ceo1: str
    ceo2: str
    nota_2025: str
    avaliacao_2025: str
    score_2025: str
    potencial_2025: str
    ceo3: str
    ceo4: str


def read_spreadsheet(path: str) -> list[dict[str, str]]:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(filename=file_path, data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        return []

    rows = worksheet.iter_rows(values_only=True)
    headers_row = next(rows, None)
    if headers_row is None:
        return []

    headers = ["" if header is None else str(header).strip() for header in headers_row]
    if not any(headers):
        return []

    parsed_rows: list[dict[str, str]] = []
    for row in rows:
        row_data: dict[str, str] = {}
        has_value = False

        for index, header in enumerate(headers):
            if header == "":
                continue

            value = row[index] if index < len(row) else None
            text_value = "" if value is None else str(value)
            row_data[header] = text_value
            if text_value != "":
                has_value = True

        if row_data and has_value:
            parsed_rows.append(row_data)

    return parsed_rows


def detect_columns(headers: list[str]) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {
        "matricula": None,
        "nome": None,
        "idade": None,
        "cargo": None,
        "antiguidade": None,
        "formacao": None,
        "resumo_perfil": None,
        "trajetoria": None,
        "nota_2025": None,
        "nota_2024": None,
        "nota_2023": None,
        "performance": None,
        "area": None,
        "potencial": None,
        "nota": None,
        "foto": None,
        "localizacao": None,
        "unidade_gestao": None,
        "ceo1": None,
        "ceo2": None,
        "ceo3": None,
        "ceo4": None,
    }

    for header in headers:
        normalized_header = _normalize_text(str(header))
        for field, accepted in NORMALIZED_VARIATIONS.items():
            if mapping[field] is None and normalized_header in accepted:
                mapping[field] = header
                break

    return mapping


def validate_required_columns(mapping: dict[str, str | None]) -> list[str]:
    return [field for field in ("nome", "cargo") if not mapping.get(field)]


def validate_ficha_required_columns(mapping: dict[str, str | None]) -> list[str]:
    return [field for field in FICHA_REQUIRED_FIELDS if not mapping.get(field)]


def validate_carom_required_columns(mapping: dict[str, str | None]) -> list[str]:
    return [field for field in CAROM_REQUIRED_FIELDS if not mapping.get(field)]


def resolve_ficha_schema(headers: list[str]) -> dict[str, str | None]:
    detected = detect_columns(headers)
    schema = {field: detected.get(field) for field in FICHA_BASE_FIELDS}
    schema.update(_resolve_ficha_evaluation_schema(headers))
    return schema


def resolve_carom_schema(headers: list[str]) -> dict[str, str | None]:
    detected = detect_columns(headers)
    schema = {
        field: detected.get(field)
        for field in (
            "matricula",
            "nome",
            "idade",
            "cargo",
            "formacao",
            "resumo_perfil",
            "trajetoria",
            "foto",
            "area",
            "localizacao",
            "unidade_gestao",
            "ceo1",
            "ceo2",
            "ceo3",
            "ceo4",
        )
    }
    evaluation_schema = _resolve_ficha_evaluation_schema(headers)
    for field in ("nota_2025", "avaliacao_2025", "score_2025", "potencial_2025"):
        schema[field] = evaluation_schema.get(field)
    return schema


def validate_standardized_ficha_schema(headers: list[str]) -> dict[str, str | None]:
    schema = resolve_ficha_schema(headers)
    missing_required = validate_ficha_required_columns(schema)
    if missing_required:
        joined = ", ".join(missing_required)
        raise ValueError(
            f"A planilha nao segue o schema padrao da ficha. Colunas ausentes: {joined}."
        )
    return schema


def validate_standardized_carom_schema(headers: list[str]) -> dict[str, str | None]:
    schema = resolve_carom_schema(headers)
    missing_required = validate_carom_required_columns(schema)
    if missing_required:
        joined = ", ".join(missing_required)
        raise ValueError(
            f"A planilha nao segue o schema padrao do carometro. Colunas ausentes: {joined}."
        )
    return schema


def has_expected_ficha_column_order(headers: list[str]) -> bool:
    normalized_headers = tuple(_normalize_text(header) for header in headers)
    for expected_order in EXPECTED_FICHA_COLUMN_ORDERS:
        normalized_expected = tuple(
            _normalize_text(header) for header in expected_order
        )
        if len(normalized_headers) >= len(normalized_expected) and (
            normalized_headers[: len(normalized_expected)] == normalized_expected
        ):
            return True
    return False


def extract_headers(rows: list[dict[str, str]]) -> list[str]:
    return list(rows[0].keys()) if rows else []


def _normalize_lookup_value(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_only.strip().lower().split())


def _find_header_by_alias(headers: list[str], aliases: tuple[str, ...]) -> str | None:
    accepted = {_normalize_text(alias) for alias in aliases}
    for header in headers:
        if _normalize_text(header) in accepted:
            return header
    return None


def _normalize_evaluation_value(value: str | None) -> str:
    if value is None:
        return ""
    compact = " ".join(str(value).split())
    if compact == "":
        return ""
    if compact.lower() in {"#n/a", "n/a"}:
        return ""
    return compact


def _build_ficha_display_note(
    *,
    direct_value: str,
    consolidated_value: str,
    score_value: str,
    potential_value: str,
) -> str:
    consolidated = _normalize_evaluation_value(consolidated_value)
    if consolidated:
        return consolidated

    score = _normalize_evaluation_value(score_value)
    potential = _normalize_evaluation_value(potential_value)
    if score and potential:
        return f"{score} / {potential}"
    if score:
        return score
    if potential:
        return potential
    direct = _normalize_evaluation_value(direct_value)
    if direct:
        return direct
    return ""


def _resolve_ficha_evaluation_schema(headers: list[str]) -> dict[str, str | None]:
    schema: dict[str, str | None] = {}
    for year in FICHA_EVALUATION_YEARS:
        consolidated_header = _find_header_by_alias(
            headers,
            (f"avaliacao_{year}", f"avaliacao {year}"),
        )
        detected_note_header = _find_header_by_alias(
            headers,
            (
                f"score_{year}",
                f"score {year}",
                f"nota_{year}",
                f"nota {year}",
            ),
        )
        potential_header = _find_header_by_alias(
            headers,
            (f"potencial_{year}", f"potencial {year}"),
        )
        direct_header = None
        score_header = detected_note_header
        if (
            detected_note_header
            and consolidated_header is None
            and potential_header is None
        ):
            direct_header = detected_note_header
            score_header = None

        schema[f"nota_{year}"] = direct_header
        schema[f"avaliacao_{year}"] = consolidated_header
        schema[f"score_{year}"] = score_header
        schema[f"potencial_{year}"] = potential_header

    return schema


def remap_ficha_row(
    row: dict[str, str], mapping: dict[str, str | None]
) -> FichaEmployee:
    normalized: dict[str, str] = {
        field: row.get(source_field, "") if source_field else ""
        for field, source_field in mapping.items()
        if field in FICHA_FIELDS
    }
    for field in FICHA_FIELDS:
        normalized.setdefault(field, "")

    for year in FICHA_EVALUATION_YEARS:
        normalized[f"avaliacao_{year}"] = _normalize_evaluation_value(
            normalized.get(f"avaliacao_{year}")
        )
        normalized[f"score_{year}"] = _normalize_evaluation_value(
            normalized.get(f"score_{year}")
        )
        normalized[f"potencial_{year}"] = _normalize_evaluation_value(
            normalized.get(f"potencial_{year}")
        )
        normalized[f"nota_{year}"] = _build_ficha_display_note(
            direct_value=normalized.get(f"nota_{year}", ""),
            consolidated_value=normalized.get(f"avaliacao_{year}", ""),
            score_value=normalized.get(f"score_{year}", ""),
            potential_value=normalized.get(f"potencial_{year}", ""),
        )

    return cast(FichaEmployee, normalized)


def remap_ficha_rows(
    rows: list[dict[str, str]], mapping: dict[str, str | None]
) -> list[FichaEmployee]:
    return [remap_ficha_row(row, mapping) for row in rows]


def load_standardized_ficha_rows(rows: list[dict[str, str]]) -> list[FichaEmployee]:
    if not rows:
        return []
    schema = validate_standardized_ficha_schema(extract_headers(rows))
    return remap_ficha_rows(rows, schema)


def remap_carom_row(
    row: dict[str, str], mapping: dict[str, str | None]
) -> CaromEmployee:
    normalized: dict[str, str] = {
        field: row.get(source_field, "") if source_field else ""
        for field, source_field in mapping.items()
        if field in CAROM_FIELDS
    }
    for field in CAROM_FIELDS:
        normalized.setdefault(field, "")
    normalized["avaliacao_2025"] = _normalize_evaluation_value(
        normalized.get("avaliacao_2025")
    )
    normalized["score_2025"] = _normalize_evaluation_value(normalized.get("score_2025"))
    normalized["potencial_2025"] = _normalize_evaluation_value(
        normalized.get("potencial_2025")
    )
    normalized["nota_2025"] = _build_ficha_display_note(
        direct_value=normalized.get("nota_2025", ""),
        consolidated_value=normalized.get("avaliacao_2025", ""),
        score_value=normalized.get("score_2025", ""),
        potential_value=normalized.get("potencial_2025", ""),
    )
    return cast(CaromEmployee, normalized)


def remap_carom_rows(
    rows: list[dict[str, str]], mapping: dict[str, str | None]
) -> list[CaromEmployee]:
    return [remap_carom_row(row, mapping) for row in rows]


def load_standardized_carom_rows(rows: list[dict[str, str]]) -> list[CaromEmployee]:
    if not rows:
        return []
    schema = validate_standardized_carom_schema(extract_headers(rows))
    return remap_carom_rows(rows, schema)


def validate_carom_employee(employee: CaromEmployee) -> list[str]:
    return [
        field
        for field in CAROM_REQUIRED_FIELDS
        if cast(str, employee.get(field, "")).strip() == ""
    ]


def resolve_carom_display_score_potential(employee: CaromEmployee) -> str:
    return _build_ficha_display_note(
        direct_value=employee.get("nota_2025", ""),
        consolidated_value=employee.get("avaliacao_2025", ""),
        score_value=employee.get("score_2025", ""),
        potential_value=employee.get("potencial_2025", ""),
    )


def carom_schema_has_display_score(mapping: dict[str, str | None]) -> bool:
    return any(
        mapping.get(field)
        for field in ("nota_2025", "avaliacao_2025", "score_2025", "potencial_2025")
    )


def validate_carom_schema_for_preset(
    mapping: dict[str, str | None],
    preset_id: str,
) -> list[str]:
    preset = get_carom_preset(preset_id)
    missing = [field for field in preset.required_fields if not mapping.get(field)]
    if preset.requires_display_score and not carom_schema_has_display_score(mapping):
        missing.extend(("nota_2025", "avaliacao_2025", "score_2025", "potencial_2025"))
    return list(dict.fromkeys(missing))


def validate_carom_employee_for_preset(
    employee: CaromEmployee,
    preset_id: str,
) -> list[str]:
    preset = get_carom_preset(preset_id)
    missing = [
        field
        for field in preset.required_fields
        if cast(str, employee.get(field, "")).strip() == ""
    ]
    if (
        preset.requires_display_score
        and resolve_carom_display_score_potential(employee) == ""
    ):
        missing.extend(("nota_2025", "avaliacao_2025", "score_2025", "potencial_2025"))
    return list(dict.fromkeys(missing))


def carom_employee_key(employee: CaromEmployee) -> str:
    matricula = employee.get("matricula", "").strip()
    if matricula:
        return f"matricula:{_normalize_lookup_value(matricula)}"
    name = _normalize_lookup_value(employee.get("nome", ""))
    cargo = _normalize_lookup_value(employee.get("cargo", ""))
    return f"fallback:{name}|{cargo}"


def filter_carom_employees(
    employees: list[CaromEmployee],
    *,
    query: str,
    mode: str,
) -> list[CaromEmployee]:
    normalized_query = _normalize_lookup_value(query)
    if normalized_query == "":
        return employees

    if mode == "matricula":
        exact_matches = [
            employee
            for employee in employees
            if _normalize_lookup_value(employee.get("matricula", ""))
            == normalized_query
        ]
        if exact_matches:
            return exact_matches
        return [
            employee
            for employee in employees
            if normalized_query
            in _normalize_lookup_value(employee.get("matricula", ""))
        ]

    return [
        employee
        for employee in employees
        if normalized_query in _normalize_lookup_value(employee.get("nome", ""))
    ]


def validate_ficha_employee(employee: FichaEmployee) -> list[str]:
    return [
        field
        for field in FICHA_REQUIRED_FIELDS
        if cast(str, employee.get(field, "")).strip() == ""
    ]


def lookup_ficha_employees(
    rows: list[dict[str, str]],
    *,
    name_query: str = "",
    matricula_query: str = "",
    max_name_matches: int = MAX_FICHA_NAME_MATCHES,
) -> list[FichaEmployee]:
    normalized_name_query = _normalize_lookup_value(name_query)
    normalized_matricula_query = _normalize_lookup_value(matricula_query)
    if normalized_name_query == "" and normalized_matricula_query == "":
        raise ValueError("Informe nome ou matricula para buscar.")

    employees = load_standardized_ficha_rows(rows)
    matches = employees

    if normalized_matricula_query:
        matches = [
            employee
            for employee in matches
            if _normalize_lookup_value(employee.get("matricula", ""))
            == normalized_matricula_query
        ]
        if len(matches) > 1:
            raise ValueError(
                "A matricula informada corresponde a mais de um colaborador na planilha."
            )

    if normalized_name_query:
        matches = [
            employee
            for employee in matches
            if normalized_name_query
            in _normalize_lookup_value(employee.get("nome", ""))
        ]

    if (
        normalized_name_query
        and not normalized_matricula_query
        and len(matches) > max_name_matches
    ):
        raise ValueError(
            "Foram encontrados muitos resultados. Refine a busca pelo nome ou informe a matricula."
        )

    return matches


def parse_multiline_field(value: str) -> list[str]:
    if value == "":
        return []

    normalized = value.replace("\r\n", "\n").replace("\r", "\n")
    items: list[str] = []
    for line in normalized.split("\n"):
        for chunk in line.split(";"):
            cleaned = chunk.strip()
            if cleaned:
                items.append(cleaned)
    return items


def normalize_filename(name: str) -> str:
    if name == "":
        return ""

    normalized = unicodedata.normalize("NFKD", name)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.replace(" ", "_")


def is_remote_source(entry: str) -> bool:
    return entry.strip().lower().startswith("https://")


def convert_onedrive_link(share_url: str) -> str:
    cleaned = share_url.strip()
    if "?download=1" in cleaned:
        return cleaned
    if "?e=" in cleaned:
        return cleaned.replace("?e=", "?download=1&e=", 1)
    separator = "&" if "?" in cleaned else "?"
    return f"{cleaned}{separator}download=1"


def get_cache_file_path(source_url: str) -> Path:
    digest = hashlib.md5(source_url.encode("utf-8")).hexdigest()
    return settings.get_cache_dir() / f"{digest}.xlsx"


def cache_is_fresh(cache_path: Path, ttl_hours: int) -> bool:
    if not cache_path.exists():
        return False
    age = datetime.now(UTC) - datetime.fromtimestamp(cache_path.stat().st_mtime, UTC)
    return age <= timedelta(hours=max(ttl_hours, 0))


def download_spreadsheet(url: str, timeout: int = 15) -> bytes:
    response = requests.get(url, timeout=timeout)
    response.raise_for_status()
    return response.content


def save_downloaded_spreadsheet(content: bytes, source_url: str) -> Path:
    cache_path = get_cache_file_path(source_url)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_bytes(content)
    return cache_path


def resolve_spreadsheet_source(
    entry: str,
    *,
    cache_enabled: bool = True,
    cache_ttl_hours: int = 24,
    force_refresh: bool = False,
) -> SpreadsheetSourceResult:
    cleaned = entry.strip()
    if cleaned == "":
        raise ValueError("Spreadsheet source is empty")

    if not is_remote_source(cleaned):
        if not Path(cleaned).exists():
            raise FileNotFoundError(cleaned)
        return SpreadsheetSourceResult(
            path=cleaned,
            source_kind="local",
            is_temporary=False,
            used_cache=False,
            message="Usando planilha local.",
        )

    source_url = convert_onedrive_link(cleaned)
    cache_path = get_cache_file_path(source_url)

    if (
        cache_enabled
        and not force_refresh
        and cache_is_fresh(cache_path, cache_ttl_hours)
    ):
        return SpreadsheetSourceResult(
            path=str(cache_path),
            source_kind="onedrive",
            is_temporary=False,
            used_cache=True,
            cache_path=str(cache_path),
            message="Usando cache local recente.",
            downloaded_at=datetime.fromtimestamp(
                cache_path.stat().st_mtime, UTC
            ).isoformat(),
        )

    try:
        content = download_spreadsheet(source_url)
        if cache_enabled:
            saved_path = save_downloaded_spreadsheet(content, source_url)
            return SpreadsheetSourceResult(
                path=str(saved_path),
                source_kind="onedrive",
                is_temporary=False,
                used_cache=False,
                cache_path=str(saved_path),
                message="Planilha atualizada a partir do OneDrive.",
                downloaded_at=datetime.now(UTC).isoformat(),
            )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name
        return SpreadsheetSourceResult(
            path=temp_path,
            source_kind="onedrive",
            is_temporary=True,
            used_cache=False,
            message="Planilha baixada temporariamente do OneDrive.",
            downloaded_at=datetime.now(UTC).isoformat(),
        )
    except Exception as exc:
        if cache_enabled and cache_path.exists():
            return SpreadsheetSourceResult(
                path=str(cache_path),
                source_kind="onedrive",
                is_temporary=False,
                used_cache=True,
                cache_path=str(cache_path),
                message=f"Falha ao atualizar a base; usando cache local. {exc}",
                downloaded_at=datetime.fromtimestamp(
                    cache_path.stat().st_mtime, UTC
                ).isoformat(),
            )
        raise RuntimeError(
            "Nao foi possivel acessar o link. Verifique a conexao com a rede da organizacao."
        ) from exc


def cleanup_source(result: SpreadsheetSourceResult) -> None:
    if result.is_temporary:
        try:
            os.unlink(result.path)
        except OSError:
            pass


def _build_performance_from_annual_notes(normalized: dict[str, str]) -> str:
    items: list[str] = []
    for year in FICHA_EVALUATION_YEARS:
        value = _build_ficha_display_note(
            direct_value=normalized.get(f"nota_{year}", ""),
            consolidated_value=normalized.get(f"avaliacao_{year}", ""),
            score_value=normalized.get(f"score_{year}", ""),
            potential_value=normalized.get(f"potencial_{year}", ""),
        )
        if value:
            items.append(f"{year} - {value}")
    return "\n".join(items)


def remap_rows(
    rows: list[dict[str, str]], mapping: dict[str, str | None]
) -> list[dict[str, str]]:
    normalized_rows: list[dict[str, str]] = []
    for row in rows:
        normalized: dict[str, str] = {}
        for target_field, source_field in mapping.items():
            normalized[target_field] = row.get(source_field, "") if source_field else ""

        annual_performance = _build_performance_from_annual_notes(normalized)
        if annual_performance:
            normalized["performance"] = annual_performance

        normalized_rows.append(normalized)
    return normalized_rows


def detect_columns_from_source(path: str) -> dict[str, str | None]:
    rows = read_spreadsheet(path)
    headers = extract_headers(rows)
    return detect_columns(headers)
